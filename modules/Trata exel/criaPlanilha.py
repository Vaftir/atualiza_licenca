import openpyxl
from openpyxl.utils import get_column_letter

class CriaPlanilha:
    def __init__(self, workbook_path, sheet_name, plataforma_acesso):
        self.workbook_path = workbook_path
        self.sheet_name = sheet_name
        self.plataforma_acesso = plataforma_acesso
        self.workbook = openpyxl.load_workbook(self.workbook_path)
        self.sheet = self.workbook[self.sheet_name]

    def obter_cabecalho(self,row_number):
        return self.sheet[row_number]

    def iterar_linhas(self, min_row=1):
        return self.sheet.iter_rows(min_row=min_row)

    def salvar_planilha(self):
        self.workbook.save(self.workbook_path)

    def inserir_linha(self, data, row_number):
        self.sheet.insert_rows(row_number)
        for col, value in enumerate(data, start=1):
            cell = self.sheet.cell(row=row_number, column=col)
            cell.value = value

    def excluir_linha(self, row_number):
        self.sheet.delete_rows(row_number)

    def atualizar_celula(self, row_number, col_number, value):
        cell = self.sheet.cell(row=row_number, column=col_number)
        cell.value = value

    def fechar_planilha(self):
        self.workbook.close()
